import openpyxl
import pandas as pd # comment

df1 =openpyxl.load_workbook (r'C:\Python27\Test folder\test.xlsx')
df =openpyxl.load_workbook (r'C:\Python27\Test folder\test1.xlsx')
sheet_df=df.active
sheet_df1 = df1.active
maxRow = sheet_df.max_row
maxCol = sheet_df.max_column
for row in range (0,maxRow):
    cell_df=sheet_df.cell(row+1,column=1)
    cell_df1=sheet_df1.cell(row+1,column=1)
    if cell_df==cell_df1:
        for col in range(0,maxCol)
            row+1
        print(cell_df.value)
